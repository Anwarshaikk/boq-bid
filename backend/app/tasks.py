import os
import time
import logging
import io
import zipfile
import base64
import json
import tempfile
import subprocess
import shutil
from collections import defaultdict
from pathlib import Path

import ezdxf
import pdfplumber
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sentence_transformers import SentenceTransformer, util
from duckduckgo_search import DDGS
import requests
from bs4 import BeautifulSoup

# Import the necessary parts from ezdxf
from ezdxf import DXFStructureError
from ezdxf.recover import read as recover_read
from ezdxf.path import make_path

from rq import get_current_job
from redis import Redis

# --- AI and Model Configuration (Unchanged) ---
_embedder = None
_dar_library = None
_layer_mapping_config = None

def get_embedder():
    global _embedder
    if _embedder is None:
        _embedder = SentenceTransformer('all-MiniLM-L6-v2')
    return _embedder

def get_dar_library():
    global _dar_library
    if _dar_library is None:
        try:
            lib_path = os.path.join(os.path.dirname(__file__), 'dar_cost_library.json')
            with open(lib_path, 'r') as f:
                _dar_library = json.load(f)
        except FileNotFoundError:
            _dar_library = []
    return _dar_library

def get_gemini_model():
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY environment variable not set.")
    genai.configure(api_key=GEMINI_API_KEY)
    return genai.GenerativeModel('gemini-1.5-flash-latest')

def clean_json_response(text):
    if '```json' in text:
        text = text.split('```json', 1)[1]
    if '```' in text:
        text = text.rsplit('```', 1)[0]
    return text.strip()

# --- NEW: ODA-Based Drawing Processing ---

def get_layer_mapping_config():
    """Loads and caches the layer mapping configuration from JSON."""
    global _layer_mapping_config
    if _layer_mapping_config is None:
        try:
            config_path = os.path.join(os.path.dirname(__file__), 'layer_mapping.json')
            with open(config_path, 'r') as f:
                _layer_mapping_config = json.load(f)
        except FileNotFoundError:
            print("🔥 CRITICAL: layer_mapping.json not found.", flush=True)
            _layer_mapping_config = {"defaults": {}, "categories": {}}
    return _layer_mapping_config

def convert_dwg_to_dxf_oda(dwg_path, base_temp_dir):
    """
    Converts a DWG file to DXF using the ODAFileConverter.
    This function is self-contained and manages its own directories.
    """
    input_filename = Path(dwg_path).name
    # Create isolated input/output folders within the main temp directory
    oda_input_dir = os.path.join(base_temp_dir, "oda_in")
    oda_output_dir = os.path.join(base_temp_dir, "oda_out")
    os.makedirs(oda_input_dir, exist_ok=True)
    os.makedirs(oda_output_dir, exist_ok=True)

    # ODA converter works on directories, so copy the file into the input dir
    shutil.copy(dwg_path, oda_input_dir)

    # Command for ODAFileConverter
    # Path is /opt/oda/ODAFileConverter as per the .deb installer
    command = [
        "/opt/oda/ODAFileConverter",
        oda_input_dir,
        oda_output_dir,
        "ACAD2018", # Output version
        "DXF",      # Output format
        "0",        # Recurse subfolders (0 for no)
        "1"         # Audit (1 for yes)
    ]
    
    print(f"🔩 Converting '{input_filename}' to DXF using ODA File Converter...", flush=True)
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=120, check=True)
        
        output_filename_dxf = f"{Path(input_filename).stem}.dxf"
        converted_dxf_path = os.path.join(oda_output_dir, output_filename_dxf)

        if os.path.exists(converted_dxf_path):
            print(f"✅ ODA Conversion successful. Output at: {converted_dxf_path}", flush=True)
            # Move the final DXF back to the base temp dir for processing
            final_path = os.path.join(base_temp_dir, output_filename_dxf)
            shutil.move(converted_dxf_path, final_path)
            return final_path
        else:
            print(f"🔥 ERROR: ODA conversion command ran but output file not found.", flush=True)
            print(f"   ODA stdout: {result.stdout}", flush=True)
            return None
    except FileNotFoundError:
        print("🔥 CRITICAL: 'ODAFileConverter' not found. Check installation path in Dockerfile.", flush=True)
        raise
    except subprocess.CalledProcessError as e:
        print(f"🔥 ERROR: ODAFileConverter failed for '{input_filename}'.", flush=True)
        print(f"   Stderr: {e.stderr}", flush=True)
        return None

def _calculate_quantities(doc, config):
    """Helper to extract quantities based on layer mapping (Unchanged)."""
    msp = doc.modelspace()
    results = defaultdict(lambda: {'quantity': 0, 'unit': '', 'description': ''})
    
    layer_to_category = {}
    for category_name, details in config['categories'].items():
        for layer in details['layers']:
            layer_to_category[layer.upper()] = {
                'name': category_name,
                'description': details['description'],
                'method': details['calculation_method']
            }

    default_height = config.get('defaults', {}).get('floor_height_m', 3.0)

    for entity in msp:
        entity_layer = entity.dxf.layer.upper()
        if entity_layer not in layer_to_category:
            continue

        category = layer_to_category[entity_layer]
        method = category['method']
        description = category['description']
        
        quantity = 0
        unit = ''

        try:
            if method == 'count':
                quantity = 1
                unit = 'Nos'
            elif method == 'length_x_height':
                if hasattr(entity, 'length'):
                    quantity = entity.length * default_height
                    unit = 'sq. m'
            elif method == 'area':
                 if hasattr(entity, 'is_closed') and entity.is_closed:
                    quantity = abs(make_path(entity).area())
                    unit = 'sq. m'
            elif method == 'area_x_height':
                if hasattr(entity, 'is_closed') and entity.is_closed:
                    area = abs(make_path(entity).area())
                    quantity = area * default_height
                    unit = 'cu. m'

            if quantity > 0:
                key = (description, unit)
                results[key]['quantity'] += quantity
                results[key]['unit'] = unit
                results[key]['description'] = description
        except Exception as e:
            print(f"⚠️ Could not calculate quantity for entity on layer '{entity_layer}'. Reason: {e}", flush=True)
    return [data for data in results.values() if data['quantity'] > 0]


def process_dwg(file_content, filename):
    """
    Parses a DWG, DXF, or ZIP file using the ODA Converter strategy.
    """
    print(f"🛠️ Worker starting ODA-based quantity take-off on: {filename}", flush=True)
    config = get_layer_mapping_config()
    final_items = []

    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = os.path.join(temp_dir, filename)
        with open(input_path, 'wb') as f:
            f.write(file_content)

        files_to_process = []
        if filename.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(input_path, 'r') as zip_ref:
                    for member_name in zip_ref.namelist():
                        if member_name.lower().endswith(('.dxf', '.dwg')):
                            extracted_path = zip_ref.extract(member_name, temp_dir)
                            files_to_process.append(extracted_path)
            except zipfile.BadZipFile:
                return {'status': 'error', 'message': f"File '{filename}' is not a valid ZIP file."}
        else:
            files_to_process.append(input_path)

        for file_path in files_to_process:
            dxf_path = file_path
            # Standardize input by converting DWG to DXF using ODA
            if file_path.lower().endswith('.dwg'):
                dxf_path = convert_dwg_to_dxf_oda(file_path, temp_dir)
                if not dxf_path:
                    print(f"Skipping file {Path(file_path).name} due to ODA conversion failure.", flush=True)
                    continue

            try:
                doc, auditor = recover_read(dxf_path)
                if auditor.has_errors:
                    print(f"⚠️ Recovered '{Path(dxf_path).name}' with errors. Proceeding...", flush=True)
                
                items = _calculate_quantities(doc, config)
                final_items.extend(items)
            except (IOError, DXFStructureError) as e:
                error_message = f"File '{Path(dxf_path).name}' is invalid or could not be read."
                print(f"🔥 {error_message} Reason: {e}", flush=True)
                continue

    consolidated = defaultdict(lambda: {'quantity': 0, 'unit': '', 'description': ''})
    for item in final_items:
        key = (item['description'], item['unit'])
        consolidated[key]['quantity'] += item['quantity']
        consolidated[key]['unit'] = item['unit']
        consolidated[key]['description'] = item['description']
    
    return {"file": filename, "items": list(consolidated.values())}

# --- UNCHANGED SECTIONS (Document AI, BoQ Generation, Strategic Intelligence) ---

def process_agreement_doc(file_content, filename):
    print(f"📄 Analyzing Agreement: {filename}", flush=True)
    model = get_gemini_model()
    try:
        with io.BytesIO(file_content) as stream, pdfplumber.open(stream) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
    except Exception as e: return {"error": f"Could not read PDF: {e}"}
    prompt = "Analyze the 'Agreement Document' text. Extract required project items. Return a valid JSON array of objects, each with 'item_code' and 'description' keys.\nExample: [{\"item_code\": \"ELEC-01\", \"description\": \"Supply and install 15W LED lights\"}]\nText:\n---\n" + text + "\n---"
    try:
        response = model.generate_content(prompt)
        return {"items": json.loads(clean_json_response(response.text))}
    except Exception as e: return {"error": f"AI parsing failed for agreement: {e}", "items": []}

def process_rules_doc(file_content, filename):
    print(f"📜 Analyzing Rules: {filename}", flush=True)
    model = get_gemini_model()
    try:
        with io.BytesIO(file_content) as stream, pdfplumber.open(stream) as pdf:
            text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
    except Exception as e: return {"error": f"Could not read PDF: {e}"}
    prompt = "Analyze the 'Rules and Specifications' text. Create a valid JSON object to act as a 'Rules Engine'. Keys should be item names, and values should be objects containing specifications.\nExample: {\"LED lights\": {\"wattage\": \"15W\", \"brand\": \"Philips\"}}\nText:\n---\n" + text + "\n---"
    try:
        response = model.generate_content(prompt)
        return {"rules": json.loads(clean_json_response(response.text))}
    except Exception as e: return {"error": f"AI parsing failed for rules: {e}", "rules": {}}

def run_auto_mapping(drawing_quantities, agreement_items, rules_engine):
    print("🔗 Auto-mapping quantities to items...", flush=True)
    embedder = get_embedder()
    if not drawing_quantities or not agreement_items: return []
    drawing_descs = [item['description'] for item in drawing_quantities]
    agreement_descs = [item['description'] for item in agreement_items]
    draw_embed = embedder.encode(drawing_descs, convert_to_tensor=True)
    agree_embed = embedder.encode(agreement_descs, convert_to_tensor=True)
    scores = util.cos_sim(draw_embed, agree_embed)
    boq = []
    for i, item in enumerate(drawing_quantities):
        idx = scores[i].argmax()
        match = agreement_items[idx]
        boq.append({**item, "mapped_item_code": match.get('item_code'), "mapped_item_description": match['description'], "match_confidence": round(scores[i][idx].item(), 2)})
    return boq
    
def apply_dar_costs(mapped_boq):
    print("💰 Applying DAR costs...", flush=True)
    dar_library = get_dar_library()
    if not dar_library: return {"error": "DAR Cost Library is not loaded or is empty."}
    embedder = get_embedder()
    costed_boq = []
    dar_descs = [item['description'] for item in dar_library]
    dar_embeddings = embedder.encode(dar_descs, convert_to_tensor=True)
    boq_descs = [item['mapped_item_description'] for item in mapped_boq]
    boq_embeddings = embedder.encode(boq_descs, convert_to_tensor=True)
    cosine_scores = util.cos_sim(boq_embeddings, dar_embeddings)
    for i, item in enumerate(mapped_boq):
        dar_match_index = cosine_scores[i].argmax()
        dar_item = dar_library[dar_match_index]
        unit_rate = dar_item.get('rate', 0.0)
        quantity = item.get('quantity', 0.0)
        costed_item = {**item, 'unit_rate': unit_rate, 'total_cost': unit_rate * quantity, 'dar_code': dar_item.get('code', 'N/A')}
        costed_boq.append(costed_item)
    excel_bytes = generate_excel_export(costed_boq)
    excel_b64 = base64.b64encode(excel_bytes).decode('utf-8')
    return {"costed_boq": costed_boq, "excel_data_b64": excel_b64}
    
def generate_excel_export(costed_boq):
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Bill of Quantities"
    headers = ["DAR Code", "Item Description", "Quantity", "Unit", "Unit Rate (INR)", "Total Cost (INR)"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center")
    for item in costed_boq:
        ws.append([item.get('dar_code'), item.get('mapped_item_description'), item.get('quantity'), item.get('unit'), item.get('unit_rate'), item.get('total_cost')])
    grand_total = sum(item.get('total_cost', 0) for item in costed_boq)
    ws.append([]); ws.append(["", "", "", "", "Grand Total", grand_total])
    total_font = Font(bold=True, size=14)
    ws.cell(row=ws.max_row, column=5).font = total_font
    ws.cell(row=ws.max_row, column=6).font = total_font
    buffer = io.BytesIO(); wb.save(buffer)
    return buffer.getvalue()

def discover_and_extract_tenders():
    print("🚀 Starting AI Tender Discovery Pipeline...", flush=True)
    all_tenders = []
    search_queries = ["latest construction tenders India", "new infrastructure projects India government", "central public works department tenders", "state PWD tenders India"]
    discovered_urls = set()
    try:
        with DDGS() as ddgs:
            for query in search_queries:
                for r in ddgs.text(query, max_results=3): discovered_urls.add(r['href'])
    except Exception as e: return {"error": "Could not perform web search for tender sources."}
    model = get_gemini_model()
    for url in list(discovered_urls)[:5]:
        try:
            response = requests.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            page_text = soup.get_text(separator=' ', strip=True)
            prompt = f"""From the text of {url}, extract active construction projects. Provide a JSON array of objects with keys: "projectName", "projectBudget", "projectLocation", "submissionDeadline". If a value is missing, use null. If no tenders, return []. Text: --- {page_text[:10000]} ---"""
            ai_response = model.generate_content(prompt)
            extracted_data = json.loads(clean_json_response(ai_response.text))
            if isinstance(extracted_data, list) and extracted_data:
                for item in extracted_data: item['sourceUrl'] = url
                all_tenders.extend(extracted_data)
        except Exception: continue
    return { "tenders": all_tenders }
    
def analyze_competitor_from_web(competitor_name):
    print(f"🧠 Analyzing competitor: {competitor_name}", flush=True)
    context = ""
    try:
        with DDGS() as ddgs:
            for r in ddgs.text(f'"{competitor_name}" construction India news', max_results=5):
                context += f"Source: {r['title']}\nSnippet: {r['body']}\n\n"
        if not context: return {"error": f"Could not find recent news for {competitor_name}."}
        model = get_gemini_model()
        prompt = f"""Analyze the news about '{competitor_name}'. Provide a concise strategic analysis in a JSON object with keys: "competitor", "projects_analyzed", "win_rate", "avg_margin", "insight".\n\nNews:\n{context}"""
        response = model.generate_content(prompt)
        return json.loads(clean_json_response(response.text))
    except Exception as e: return {"error": "AI analysis failed."}
